# -*- coding: utf-8 -*-
"""Annotate C4 haplotype calls with all-public-sample region metadata.

This script parameterizes the final C4 strict-filter support step.
It adds region-related columns to the ``haplotype_calls`` sheet while keeping
the original workbook unchanged by default.
"""

from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ANNOTATION_COLUMNS = [
    "sample_id_region_key",
    "superpopulation_code",
    "population_code",
    "region_annotation_source",
    "keep_public_reference_east_asian_subset",
]

SAMPLE_PATTERNS = [
    r"HG\d+",
    r"GM\d+",
    r"NA\d+",
    r"CN1",
    r"HG005",
    r"KOREF1",
    r"YAO",
    r"CHM13",
    r"KSA\d+",
    r"C\d{3}-CHA-E\d+",
    r"AK\d+",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Annotate C4 haplotype_calls rows with population-region metadata."
    )
    parser.add_argument("input_xlsx", type=Path, help="Strict C4 result workbook.")
    parser.add_argument(
        "--mapping-tsv",
        type=Path,
        required=True,
        help="public_reference sample-to-region mapping table.",
    )
    parser.add_argument(
        "--eas-sample-ids",
        type=Path,
        required=True,
        help="Plain-text list of public_reference sample IDs retained as EAS.",
    )
    parser.add_argument(
        "-o",
        "--output-xlsx",
        type=Path,
        help="Annotated workbook. Defaults to <input stem>_region_annotated.xlsx.",
    )
    parser.add_argument(
        "--sheet",
        default="haplotype_calls",
        help="Worksheet to annotate.",
    )
    parser.add_argument(
        "--preview-rows",
        type=int,
        default=300,
        help="Maximum number of preview rows to write.",
    )
    return parser.parse_args()


def require_file(path: Path) -> None:
    if not path.is_file():
        raise FileNotFoundError(f"Required file not found: {path}")


def load_region_mapping(
    mapping_tsv: Path,
    eas_sample_ids: Path,
) -> tuple[dict[str, dict[str, str]], set[str]]:
    mapping = pd.read_csv(mapping_tsv, sep="\t", dtype=str).fillna("")
    eas_ids = {
        line.strip()
        for line in eas_sample_ids.read_text(encoding="utf-8").splitlines()
        if line.strip()
    }

    region_by_sample: dict[str, dict[str, str]] = {}
    for _, row in mapping.iterrows():
        sample_id = str(row.get("sample_id", "")).strip()
        if not sample_id:
            continue
        region_by_sample[sample_id] = {
            "region": str(row.get("superpopulation_code", "")).strip(),
            "population": str(row.get("population_code", "")).strip(),
            "source": str(row.get("source_project", "")).strip(),
            "keep": str(row.get("keep_EAS", "")).strip(),
            "status": str(row.get("annotation_status", "")).strip(),
        }

    for sample_id in eas_ids:
        region_by_sample.setdefault(
            sample_id,
            {
                "region": "EAS",
                "population": "",
                "source": "EAS_sample_ids.txt",
                "keep": "yes",
                "status": "EAS_list",
            },
        )
        region_by_sample[sample_id]["region"] = (
            region_by_sample[sample_id].get("region") or "EAS"
        )
        region_by_sample[sample_id]["keep"] = (
            region_by_sample[sample_id].get("keep") or "yes"
        )
    return region_by_sample, eas_ids


def build_sample_extractor(known_ids: list[str]):
    sample_re = re.compile("|".join(f"({pattern})" for pattern in SAMPLE_PATTERNS))

    def extract_sample_id(value: object) -> str:
        if value is None:
            return ""
        text = str(value)
        text = re.sub(r"^(C4AL|C4AS|C4BL|C4BS|HERV)_", "", text)
        for sample_id in known_ids:
            if sample_id and sample_id in text:
                return sample_id
        match = re.search(r"\d+_([A-Za-z0-9-]+)\.(pat|mat|pri)", text)
        if match:
            return match.group(1)
        match = re.match(r"([A-Z]{2}\d{5}|HG\d{5}|NA\d{5})[._-]", text)
        if match:
            return match.group(1)
        match = sample_re.search(text)
        if match:
            return next(group for group in match.groups() if group)
        return ""

    return extract_sample_id


def infer_local_dataset_region(dataset: object) -> tuple[str, str, str, str] | None:
    dataset_text = str(dataset).upper()
    if "SCZ" in dataset_text:
        return "EAS", "SCZ", "dataset_label_local_EAS", "no"
    if "comparison_cohort" in dataset_text or "comparison_site" in dataset_text:
        return "EAS", "comparison_cohort", "dataset_label_local_EAS", "yes"
    return None


def remove_existing_annotation_columns(worksheet) -> None:
    existing = {
        str(worksheet.cell(row=1, column=i).value): i
        for i in range(1, worksheet.max_column + 1)
    }
    for header in reversed(ANNOTATION_COLUMNS):
        if header in existing:
            worksheet.delete_cols(existing[header], 1)
            existing = {
                str(worksheet.cell(row=1, column=i).value): i
                for i in range(1, worksheet.max_column + 1)
            }


def annotate_workbook(
    input_xlsx: Path,
    output_xlsx: Path,
    mapping_tsv: Path,
    eas_sample_ids: Path,
    sheet_name: str,
    preview_rows: int,
) -> None:
    region_by_sample, eas_ids = load_region_mapping(mapping_tsv, eas_sample_ids)
    known_ids = sorted(region_by_sample, key=len, reverse=True)
    extract_sample_id = build_sample_extractor(known_ids)

    workbook = load_workbook(input_xlsx)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"No {sheet_name!r} sheet in {input_xlsx}; sheets={workbook.sheetnames}")
    worksheet = workbook[sheet_name]
    headers = [
        "" if worksheet.cell(1, i).value is None else str(worksheet.cell(1, i).value)
        for i in range(1, worksheet.max_column + 1)
    ]
    sample_col = headers.index("sample_id") + 1 if "sample_id" in headers else 2
    dataset_col = headers.index("dataset") + 1 if "dataset" in headers else None

    remove_existing_annotation_columns(worksheet)
    start_col = worksheet.max_column + 1
    for offset, header in enumerate(ANNOTATION_COLUMNS):
        worksheet.cell(row=1, column=start_col + offset).value = header

    counts: Counter[str] = Counter()
    by_dataset: defaultdict[str, Counter[str]] = defaultdict(Counter)
    preview = []

    for row_idx in range(2, worksheet.max_row + 1):
        raw_sample = worksheet.cell(row=row_idx, column=sample_col).value
        dataset = (
            str(worksheet.cell(row=row_idx, column=dataset_col).value)
            if dataset_col
            else ""
        )
        sample_id = extract_sample_id(raw_sample)
        info = region_by_sample.get(sample_id)
        local_region = infer_local_dataset_region(dataset)

        if info:
            region = info.get("region", "") or "unresolved"
            population = info.get("population", "")
            source = info.get("source", "public_reference_sample_region_mapping.tsv")
            keep = (
                "yes"
                if sample_id in eas_ids
                or region.upper() == "EAS"
                or str(info.get("keep", "")).lower() == "yes"
                else "no"
            )
        elif local_region:
            region, population, source, keep = local_region
        elif not sample_id:
            region = "unparsed"
            population = ""
            source = "sample_id_parse_failed"
            keep = ""
        else:
            region = "unresolved_or_non_public_reference"
            population = ""
            source = "not_found_in_public_reference_mapping"
            keep = "no"

        values = [sample_id, region, population, source, keep]
        for offset, value in enumerate(values):
            worksheet.cell(row=row_idx, column=start_col + offset).value = value

        counts[region] += 1
        by_dataset[dataset][region] += 1
        if len(preview) < preview_rows:
            preview.append(
                {
                    "dataset": dataset,
                    "sample_id_original": raw_sample,
                    "sample_id_region_key": sample_id,
                    "superpopulation_code": region,
                    "population_code": population,
                    "region_annotation_source": source,
                    "keep_public_reference_east_asian_subset": keep,
                }
            )

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)

    summary_path = output_xlsx.with_name(output_xlsx.stem + "_region_annotation_summary.tsv")
    preview_path = output_xlsx.with_name(output_xlsx.stem + "_region_annotation_preview.tsv")
    summary_rows = [
        ("source_xlsx", str(input_xlsx)),
        ("annotated_xlsx", str(output_xlsx)),
        ("sheet", sheet_name),
        ("rows_annotated", str(worksheet.max_row - 1)),
        ("sample_id_column", headers[sample_col - 1]),
        ("public_reference_mapping", str(mapping_tsv)),
        ("eas_sample_ids", str(len(eas_ids))),
    ]
    for region, count in counts.most_common():
        summary_rows.append((f"region_count:{region}", str(count)))
    for dataset, counter in sorted(by_dataset.items()):
        for region, count in counter.most_common():
            summary_rows.append((f"dataset_region_count:{dataset}:{region}", str(count)))

    summary_path.write_text(
        "metric\tvalue\n"
        + "\n".join(f"{key}\t{value}" for key, value in summary_rows)
        + "\n",
        encoding="utf-8",
    )
    pd.DataFrame(preview).to_csv(preview_path, sep="\t", index=False, encoding="utf-8")

    print(f"Saved: {output_xlsx}")
    print(f"Saved: {summary_path}")
    print(f"Saved: {preview_path}")
    print(f"Region counts: {dict(counts)}")


def main() -> None:
    args = parse_args()
    require_file(args.input_xlsx)
    require_file(args.mapping_tsv)
    require_file(args.eas_sample_ids)

    output_xlsx = args.output_xlsx or args.input_xlsx.with_name(
        f"{args.input_xlsx.stem}_region_annotated.xlsx"
    )
    annotate_workbook(
        input_xlsx=args.input_xlsx,
        output_xlsx=output_xlsx,
        mapping_tsv=args.mapping_tsv,
        eas_sample_ids=args.eas_sample_ids,
        sheet_name=args.sheet,
        preview_rows=args.preview_rows,
    )


if __name__ == "__main__":
    main()
