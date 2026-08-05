# -*- coding: utf-8 -*-
"""Verify C4 haplotype-call region annotation columns."""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_COLUMNS = [
    "sample_id_region_key",
    "superpopulation_code",
    "population_code",
    "region_annotation_source",
    "keep_public_reference_east_asian_subset",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Summarize region annotations in a C4 haplotype_calls sheet."
    )
    parser.add_argument("input_xlsx", type=Path, help="Annotated C4 workbook.")
    parser.add_argument(
        "--sheet",
        default="haplotype_calls",
        help="Worksheet to verify.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.input_xlsx.is_file():
        raise FileNotFoundError(f"Input workbook not found: {args.input_xlsx}")

    workbook = load_workbook(args.input_xlsx, read_only=True, data_only=False)
    if args.sheet not in workbook.sheetnames:
        raise ValueError(f"No {args.sheet!r} sheet in {args.input_xlsx}")
    worksheet = workbook[args.sheet]

    headers = [
        "" if worksheet.cell(1, i).value is None else str(worksheet.cell(1, i).value)
        for i in range(1, worksheet.max_column + 1)
    ]
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise RuntimeError(f"Missing annotation columns: {missing}")

    region_col = headers.index("superpopulation_code") + 1
    dataset_col = headers.index("dataset") + 1 if "dataset" in headers else None
    counts: Counter[str] = Counter()
    by_dataset: defaultdict[str, Counter[str]] = defaultdict(Counter)

    for row_idx in range(2, worksheet.max_row + 1):
        region = worksheet.cell(row_idx, region_col).value
        region = "" if region is None else str(region)
        counts[region] += 1
        if dataset_col:
            dataset = worksheet.cell(row_idx, dataset_col).value
            by_dataset[str(dataset)][region] += 1

    print(f"Workbook: {args.input_xlsx}")
    print(f"Rows: {worksheet.max_row - 1}")
    print(f"Annotation columns: {REQUIRED_COLUMNS}")
    print(f"Region counts: {dict(counts)}")
    if dataset_col:
        print(f"Dataset-region counts: { {key: dict(value) for key, value in by_dataset.items()} }")


if __name__ == "__main__":
    main()
