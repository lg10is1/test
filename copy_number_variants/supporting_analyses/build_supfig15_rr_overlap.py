# -*- coding: utf-8 -*-

from __future__ import annotations

import gzip
import os
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_DIR = SCRIPT_DIR / "inputs"

SCZ_GENE_TABLE = Path(
    os.environ.get(
        "EOSCZ_SUPFIG15_SCZ_GENE_TABLE",
        INPUT_DIR / "gene_frequencies_with_chr_length_case_cohort.txt",
    )
)
REPORTED_CNV_TABLE = Path(
    os.environ.get(
        "EOSCZ_SUPFIG15_REPORTED_CNV_TABLE",
        INPUT_DIR / "41583_2024_837_MOESM1_ESM.xlsx",
    )
)
HG38_TO_CHM13_CHAIN = Path(
    os.environ.get(
        "EOSCZ_HG38_TO_CHM13_CHAIN",
        INPUT_DIR / "hg38-chm13v2.over.chain.gz",
    )
)

OUTPUT_XLSX = SCRIPT_DIR / "Supplementary_Figure_15_seven_genes_overlapped_with_RR_updated_26-7-14.xlsx"
OUTPUT_TSV = SCRIPT_DIR / "Supplementary_Figure_15_seven_genes_overlapped_with_RR_updated_26-7-14.tsv"

FINAL_GENE_ORDER = ["TBX6", "ZG16", "ASPHD1", "GDPD3", "SLC25A1", "EIF4H", "RFC2"]


def lift_interval_to_chm13(chrom: str, start: int, end: int) -> tuple[str, int, int]:
    pieces: list[tuple[str, int, int]] = []

    with gzip.open(HG38_TO_CHM13_CHAIN, "rt") as handle:
        tname = qname = qstrand = None
        qsize = tpos = qpos = None

        for raw_line in handle:
            line = raw_line.strip()
            if not line:
                continue

            if line.startswith("chain"):
                parts = line.split()
                tname = parts[2]
                tstart = int(parts[5])
                qname = parts[7]
                qsize = int(parts[8])
                qstrand = parts[9]
                qstart = int(parts[10])
                tpos = tstart
                qpos = qstart
                continue

            values = line.split()
            block_size = int(values[0])

            if tname == chrom:
                block_start = int(tpos)
                block_end = block_start + block_size
                overlap_start = max(start, block_start)
                overlap_end = min(end, block_end)

                if overlap_start < overlap_end:
                    offset_start = overlap_start - block_start
                    offset_end = overlap_end - block_start
                    if qstrand == "+":
                        query_start = int(qpos) + offset_start
                        query_end = int(qpos) + offset_end
                    else:
                        query_start = int(qsize) - (int(qpos) + offset_end)
                        query_end = int(qsize) - (int(qpos) + offset_start)
                    pieces.append((str(qname), query_start, query_end))

            if len(values) == 3:
                _, dt, dq = map(int, values)
                tpos = int(tpos) + block_size + dt
                qpos = int(qpos) + block_size + dq
            else:
                tpos = int(tpos) + block_size
                qpos = int(qpos) + block_size

    if not pieces:
        raise ValueError(f"No CHM13 liftover result for {chrom}:{start}-{end}")

    by_chrom: dict[str, list[tuple[int, int]]] = defaultdict(list)
    for out_chrom, out_start, out_end in pieces:
        by_chrom[out_chrom].append((out_start, out_end))

    if len(by_chrom) != 1:
        raise ValueError(f"Liftover mapped {chrom}:{start}-{end} to multiple chromosomes: {sorted(by_chrom)}")

    out_chrom, intervals = next(iter(by_chrom.items()))
    return out_chrom, min(start for start, _ in intervals), max(end for _, end in intervals)


def main() -> None:
    scz_genes = pd.read_table(SCZ_GENE_TABLE)
    scz_genes = scz_genes.rename(
        columns={
            "Chromosome": "Gene \nChm13 \nChr",
            "Start": "Gene \nChm13 \nStart",
            "End": "Gene \nChm13 \nEnd",
        }
    )

    reported = pd.read_excel(REPORTED_CNV_TABLE, sheet_name="tableS3")
    reported = reported.loc[reported["type"].eq("dup")].copy()

    reported[["RR \nChm13 \nChr", "RR \nChm13 \nStart", "RR \nChm13 \nEnd"]] = reported.apply(
        lambda row: pd.Series(lift_interval_to_chm13(row["hg38chr"], int(row["start"]), int(row["end"]))),
        axis=1,
    )

    hits = []
    for _, gene in scz_genes.iterrows():
        for _, rr in reported.iterrows():
            if (
                gene["Gene \nChm13 \nChr"] == rr["RR \nChm13 \nChr"]
                and int(gene["Gene \nChm13 \nStart"]) < int(rr["RR \nChm13 \nEnd"])
                and int(gene["Gene \nChm13 \nEnd"]) > int(rr["RR \nChm13 \nStart"])
            ):
                hits.append(
                    {
                        "Gene": gene["Gene"],
                        "Gene \nChm13 \nChr": gene["Gene \nChm13 \nChr"],
                        "Gene \nChm13 \nStart": int(gene["Gene \nChm13 \nStart"]),
                        "Gene \nChm13 \nEnd": int(gene["Gene \nChm13 \nEnd"]),
                        "RR \nChm13 \nChr": rr["RR \nChm13 \nChr"],
                        "RR \nChm13 \nStart": int(rr["RR \nChm13 \nStart"]),
                        "RR \nChm13 \nEnd": int(rr["RR \nChm13 \nEnd"]),
                        "RR \nhg38 \nChr": rr["hg38chr"],
                        "RR \nhg38 \nStart": int(rr["start"]),
                        "RR \nhg38 \nEnd": int(rr["end"]),
                        "RR \nType": rr["type"],
                    }
                )

    out = pd.DataFrame(hits)
    out = out.loc[out["Gene"].isin(FINAL_GENE_ORDER)].copy()
    out["Gene_order"] = out["Gene"].map({gene: i for i, gene in enumerate(FINAL_GENE_ORDER)})
    out = out.sort_values("Gene_order").drop(columns="Gene_order")

    if out["Gene"].tolist() != FINAL_GENE_ORDER:
        raise ValueError(f"Unexpected gene list: {out['Gene'].tolist()}")

    out.to_csv(OUTPUT_TSV, sep="\t", index=False)

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="SupFig15_updated", index=False)
        reported.to_excel(writer, sheet_name="reported_CNV_regions_dup", index=False)

        workbook = writer.book
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            header_fill = PatternFill("solid", fgColor="D9EAF7")
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for idx, column_cells in enumerate(ws.columns, start=1):
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                ws.column_dimensions[get_column_letter(idx)].width = min(max(max_length + 2, 12), 24)
            ws.freeze_panes = "A2"

        readme = workbook.create_sheet("README")
        readme["A1"] = "Caption"
        readme["B1"] = (
            "Supplementary Figure 15. Seven genes overlapped with previously reported "
            "schizophrenia-associated CNV regions in eoSCZ samples. RR means reported region."
        )
        readme["A2"] = "SCZ gene source"
        readme["B2"] = str(SCZ_GENE_TABLE)
        readme["A3"] = "Reported CNV source"
        readme["B3"] = "41583_2024_837_MOESM1_ESM.xlsx, tableS3, type == dup"
        readme["A4"] = "Liftover"
        readme["B4"] = "hg38 reported regions were lifted to CHM13 with hg38-chm13v2.over.chain.gz; CHM13 spans use min/max of lifted blocks."
        for row in readme.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        readme.column_dimensions["A"].width = 22
        readme.column_dimensions["B"].width = 120

    print(out.to_string(index=False))
    print(f"Saved: {OUTPUT_XLSX}")
    print(f"Saved: {OUTPUT_TSV}")


if __name__ == "__main__":
    main()
